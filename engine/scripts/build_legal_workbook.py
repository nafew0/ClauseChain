"""Build the legal-review workbook (xlsx) from CURRENT artifacts — reproducibly.

Replaces the 15-Jul one-off. Sheets (layout matches the reviewed original, with
two added refuter columns on NEW Findings):
  Instructions · Recall Misses · NEW Findings · Indicator Criteria · Master Known

Sources: outputs/final_*_p*/output.{csv,json}, data/review/recall_adjudication.json,
data/review/refutation_final_*.json, configs/rdtii/pillar_{6,7}.yaml,
data/known_index.json.

Usage: .venv/bin/python scripts/build_legal_workbook.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

RUNS = ["final_si_p6", "final_si_p7", "final_ma_p6", "final_ma_p7",
        "final_au_p6", "final_au_p7"]
OUT = Path("outputs/legal_recall_review/ClauseChain_Legal_Review_Workbook.xlsx")

HEAD_FONT = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="0FB5A7")
WRAP = Alignment(wrap_text=True, vertical="top")

REVIEW_COLS = ["Reviewer decision", "Reviewer correction/reasoning",
               "Reviewer official source URL", "Reviewer name", "Reviewer role",
               "Review date", "Citation checked", "Mapping checked", "Currentness checked"]


def _sheet(wb, title, headers, widths):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"
    return ws


def _indicator_questions() -> dict[str, dict]:
    out = {}
    for p in ("6", "7"):
        cfg = yaml.safe_load(Path(f"configs/rdtii/pillar_{p}.yaml").read_text())
        for ind_id, ind in (cfg.get("indicators") or {}).items():
            out[ind_id] = ind if isinstance(ind, dict) else {}
    return out


def _alignment_label(finding: dict) -> str:
    proof = finding.get("citation_proof") or {}
    method = str(proof.get("alignment_status") or proof.get("method")
                 or proof.get("alignment") or "").lower()
    score = proof.get("alignment_score")
    if "anchor" in method or (finding.get("Location Reference") or "").startswith("#"):
        return f"anchor ({score if score is not None else 1})"
    if "exact" in method:
        return f"exact ({score if score is not None else 1})"
    return "unaligned (0)"


def _display(value, limit: int) -> str:
    if value is None:
        return ""
    text = value if isinstance(value, str) else json.dumps(value, ensure_ascii=False)
    return text[:limit]


def _surrounding_context(finding: dict) -> str:
    return _display(finding.get("raw_context") or finding.get("Notes") or "", 1400)


def _official_urls(entry: dict) -> str:
    values = entry.get("references") or entry.get("urls") or entry.get("url") or []
    if isinstance(values, str):
        values = [values]
    return "\n".join(str(value).strip() for value in values if str(value).strip())[:900]


def _recall_rationale(miss: dict) -> str:
    evidence = miss.get("evidence") or {}
    technical = evidence.get("technical_class") or miss.get("class") or "UNCLASSIFIED"
    emitted = ", ".join(miss.get("emitted_under") or [])
    messages = {
        "NOT_IN_CORPUS": (
            "The master citation was not located in the rebuilt eligible corpus. Review whether "
            "the official instrument/version was acquired and whether the master reference is current."
        ),
        "IN_CORPUS_NOT_EMITTED": (
            "The cited provision exists in the corpus, but no surviving finding was emitted for the "
            "master indicator. Review legal mapping, context and deterministic gate outcomes."
        ),
        "EMITTED_OTHER_INDICATOR": (
            f"The cited provision was emitted under {emitted or 'another indicator'}, not under the "
            "master indicator. Decide whether the master mapping, engine mapping, or both need correction."
        ),
        "GOLD_REF_UNPARSEABLE": (
            "The master citation could not be parsed into a stable legal anchor. Verify the citation "
            "against the official instrument and provide a corrected reference if necessary."
        ),
    }
    detail = messages.get(technical, f"Technical classification: {technical}. Legal review is required.")
    return f"{detail} Proposed system verdict: {miss.get('proposed_verdict') or 'REVIEW_REQUIRED'}."


def main() -> int:
    indicators = _indicator_questions()
    refutations: dict[tuple, dict] = {}
    for run in RUNS:
        path = Path(f"data/review/refutation_{run}.json")
        if not path.is_file():
            continue
        data = json.loads(path.read_text())
        items = data if isinstance(data, list) else data.get("rows", data.get("findings", []))
        for it in items:
            key = (it.get("indicator"), it.get("law"), it.get("article"))
            votes = it.get("refuter_votes") or []
            reason = " || ".join(
                f"{v.get('persona')}/{v.get('failure_mode', 'none')}: {v.get('reason','')[:220]}"
                for v in votes)
            refutations[key] = {"verdict": it.get("verdict"), "reason": reason}

    wb = Workbook()
    wb.remove(wb.active)

    # --- Instructions --------------------------------------------------------
    ws = _sheet(wb, "Instructions", ["ClauseChain Legal Review", ""], [40, 110])
    misses = json.loads(Path("data/review/recall_adjudication.json").read_text())
    ws.append(["Generated by", "scripts/build_legal_workbook.py — regenerate any time; do not hand-edit structure"])
    ws.append(["Sheet: Recall Misses", f"{len(misses.get('misses', []))} unmatched master anchors — fill Reviewer verdict "
               "(REAL_MISS / GOLD_WRONG / GOLD_AMBIGUOUS / CORRECT_ABSTENTION). Malaysia rows: ESCAP planted deliberate "
               "errors (confirmed) — cross-check the error audit before calling anything a real miss."])
    ws.append(["Sheet: NEW Findings", "Every NEW row is pending. The refuter is one configured model applying three "
               "named analytical lenses with the complete indicator rubric; it is advisory, not three independent "
               "legal opinions. Confirm or override it; your named decision is final."])
    ws.append(["Decisions vocabulary", "approve | reject | needs-correction (name + date required; row is excluded from "
               "the final export without an explicit approve)"])
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = WRAP

    # --- Recall Misses --------------------------------------------------------
    heads = ["Miss ID", "Economy", "Indicator", "Indicator question", "Master act/instrument",
             "Master citation", "Technical class", "Emitted under", "Proposed verdict",
             "System rationale"] + ["Reviewer verdict"] + REVIEW_COLS[1:]
    ws = _sheet(wb, "Recall Misses", heads, [8, 11, 9, 40, 34, 12, 18, 12, 20, 46] + [16] * 9)
    for i, m in enumerate(misses.get("misses", []), start=1):
        ind = indicators.get(m.get("gold_indicator") or "", {})
        ws.append([f"M{i:03d}", m.get("economy"), m.get("gold_indicator"),
                   str(ind.get("question", ind.get("legal_question", ind.get("name", ""))))[:300],
                   m.get("act"), m.get("ref"), m.get("class"),
                   ", ".join(m.get("emitted_under") or []) or "—",
                   m.get("proposed_verdict"),
                   _recall_rationale(m)] + [""] * 9)

    # --- NEW Findings ----------------------------------------------------------
    heads = ["NEW ID", "Economy", "Indicator", "Indicator question", "Law/instrument",
             "Article/section", "Exact source snippet", "Surrounding context",
             "System mapping rationale", "Source URL", "Alignment", "Status evidence",
             "Gate warnings", "Refuter verdict", "Refuter reasoning",
             "System legal-review comment"] + REVIEW_COLS
    ws = _sheet(wb, "NEW Findings", heads,
                [8, 11, 9, 34, 30, 13, 52, 40, 44, 30, 12, 26, 26, 18, 40, 40] + [16] * 9)
    n = 0
    for run in RUNS:
        env = json.loads(Path(f"outputs/{run}/output.json").read_text())
        gates = env.get("gates", [])
        for f in env.get("findings", []):
            if f.get("Discovery Tag") != "NEW":
                continue
            n += 1
            cite = f.get("Article / Section", "")
            row_warns = " | ".join(f"{g['gate_id']}: {g.get('reason','')[:80]}" for g in gates
                                   if g.get("status") == "WARN" and cite and cite in str(g.get("evidence_reference", "")))
            ref = refutations.get((f.get("Indicator ID"), f.get("Law Name"), cite), {})
            ind = indicators.get(f.get("Indicator ID") or "", {})
            comment = ("Technical provenance gates passed; legal must confirm substantive indicator fit, "
                       "scope, exceptions and currentness.")
            align = _alignment_label(f)
            if align.startswith("unaligned"):
                comment = "BLOCK: citation not aligned to a canonical page/anchor; cannot enter final export until resolved. " + comment
            ws.append([f"N{n:03d}", f.get("Economy"), f.get("Indicator ID"),
                   str(ind.get("question", ind.get("legal_question", ind.get("name", ""))))[:300],
                       f.get("Law Name"), cite,
                       str(f.get("Verbatim Snippet", ""))[:900],
                       _surrounding_context(f),
                       str(f.get("Mapping Rationale", ""))[:700],
                       f.get("Source URL"), align,
                       _display(f.get("status_evidence", ""), 700), row_warns,
                       str(ref.get("verdict", "NOT RUN"))[:40],
                       str(ref.get("reason", ref.get("rationale", "")))[:400],
                       comment] + [""] * 9)
    for sheet_name in ("Recall Misses", "NEW Findings"):
        for row in wb[sheet_name].iter_rows(min_row=2):
            for c in row:
                c.alignment = WRAP

    # --- Indicator Criteria ----------------------------------------------------
    ws = _sheet(wb, "Indicator Criteria",
                ["Indicator", "Methodology no.", "Name", "Legal question", "Legal test",
                 "Scoring criteria", "Exclusions"], [10, 12, 26, 44, 44, 44, 34])
    for ind_id, ind in sorted(indicators.items()):
        ws.append([ind_id, str(ind.get("methodology_no", "")), str(ind.get("name", ""))[:120],
                   str(ind.get("question", ind.get("legal_question", "")))[:400],
                   str(ind.get("legal_test", ""))[:400],
                   json.dumps(ind.get("criteria", ind.get("scoring", "")))[:400],
                   json.dumps(ind.get("exclusions", ""))[:300]])

    # --- Master Known ----------------------------------------------------------
    ws = _sheet(wb, "Master Known",
                ["Economy", "Indicator", "Methodology score", "Act/instrument",
                 "Article references", "Coverage", "Master impact/rationale", "Official URL(s)"],
                [11, 9, 10, 36, 22, 12, 60, 40])
    idx = json.loads(Path("data/known_index.json").read_text())["economies"]
    for economy in ("Singapore", "Malaysia", "Australia"):
        for e in idx.get(economy, []):
            if e.get("source") != "master" or not str(e.get("indicator_code", "")).startswith(("P6", "P7")):
                continue
            ws.append([economy, e.get("indicator_code"), str(e.get("score", "")),
                       str(e.get("act", ""))[:200], ", ".join(e.get("articles", []))[:150],
                       str(e.get("coverage", "")), str(e.get("impact", ""))[:800],
                       _official_urls(e)])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT} — NEW rows: {n}, misses: {len(misses.get('misses', []))}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
